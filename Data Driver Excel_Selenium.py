
#========Excel operation-====================
"""
Reading Data from excel file
writing data into excel file
data driven test case

"""
import openpyxl

path="C:\\Users\kandeepx\Documents\B\MM.xlsx"

workbook=openpyxl.load_workbook(path)
sheet=workbook.get_sheet_by_name("High") #excel will have multiple sheets so to specify the sheet we r using this command or else we can use "active"
rows=sheet.max_row
cols=sheet.max_column

print("High sheet has below rows and cols :")
print(rows)
print(cols)


for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value,end="   ")

    print()


#==========Writing Excel==========
"""
path="C:\\Users\kandeepx\Documents\B\\test.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.active

for r in range(1,7):
    for c in range(1,5):
        sheet.cell(row=r,column=c).value="welcome"

workbook.save(path)
"""



